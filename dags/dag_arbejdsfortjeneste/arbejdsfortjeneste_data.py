import logging
import io
import pandas as pd
from typing import Any, Iterable, Iterator, Sequence
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from rkdigi.email_handling import EmailReader
from airflow.models import Variable

logger = logging.getLogger(__name__)


def _get_arbejdsfortjeneste_report_config() -> dict[str, Any]:
    """
    Load Arbejdsfortjeneste report configuration from Airflow Variables.

    :return: Parsed JSON configuration used for extraction and change reporting.
    """
    return Variable.get("arbejdsfortjeneste_report_config", deserialize_json=True)


def _get_change_report_numeric_fields() -> list[str]:
    """
    Get numeric fields used for month-to-month change calculations.

    :return: List of numeric field names for change aggregation.
    """
    return _get_arbejdsfortjeneste_report_config().get("change_report_numeric_fields", [])


def _get_required_fields_from_blanket_16001() -> list[str]:
    """
    Get required fields from blanket 16001 used to validate extracted rows.

    :return: List of required field names from blanket 16001.
    """
    return _get_arbejdsfortjeneste_report_config().get("required_fields_from_blanket_16001", [])


def get_report_field_to_blanket_field_id() -> dict[str, str]:
    """
    Get mapping from report field names to blanket field identifiers.

    :return: Dictionary mapping report column name to blanket field id.
    """
    return _get_arbejdsfortjeneste_report_config().get("report_field_to_blanket_field_id", {})


def _get_income_type_code_to_label() -> dict[str, str]:
    """
    Get mapping from income type code to readable label.

    :return: Dictionary mapping income type code to label.
    """
    return _get_arbejdsfortjeneste_report_config().get("income_type_code_to_label", {})


def get_change_report_key_columns() -> list[str]:
    """
    Get key columns used to group and compare monthly aggregates.

    :return: List of key columns. Defaults to ["cpr"] if not configured.
    """
    return _get_arbejdsfortjeneste_report_config().get("change_report_key_columns", ["cpr"])


def map_indkomsttype(value: str | int | None) -> str | None:
    """
    Map an IndkomstType code to a configured label.

    :param value: Raw income type code from Serviceplatform SKAT payload.
    :return: Mapped label if configured
    """
    if value is None:
        return None
    key = str(value).strip()
    return _get_income_type_code_to_label().get(key, value)


def _normalize_cpr(value: object) -> str | None:
    """
    Normalize CPR value by removing - separator
    """
    if value is None or pd.isna(value):
        return None

    normalized = str(value).strip().replace("-", "")

    return normalized


def extract_cprs(
    df: pd.DataFrame,
    column: str = None,
) -> list[str]:
    """
    Extract CPR numbers from a DataFrame column without deduplicating.

    :param df: DataFrame from Excel.
    :param column: Column name containing CPR numbers.
    :return: CPR list
    """
    if column is None:
        raise ValueError("Column name must be provided for CPR extraction")

    cprs = [
        cpr
        for cpr in df[column].map(_normalize_cpr).tolist()
        if cpr is not None
    ]
    logger.info(f"Extracted {len(cprs)} CPR values from Excel column '{column}'")
    return cprs


def find_latest_attachment(
    email_reader: EmailReader,
    mailbox: str = "INBOX",
    criteria: str = "UNSEEN",
    extensions: Sequence[str] = (".xlsx",),
    filename_prefixes: Iterable[str] | None = "Liste",
    max_emails: int = 50,
) -> tuple[bytes, str, bytes] | None:
    """
    Find newest attachment matching extension and filename prefixes.

    :param email_reader: EmailReader instance used to fetch emails.
    :param mailbox: Mailbox/folder name to search in.
    :param criteria: IMAP search criteria (for example ``UNSEEN`` or ``ALL``).
    :param extensions: Allowed file extensions (case-insensitive).
    :param filename_prefixes: Optional allowed filename prefixes.
    :param max_emails: Maximum number of emails to inspect.
    :return: Tuple ``(uid, filename, content_bytes)`` or ``None`` if no match.
    """
    emails, failed = email_reader.get_emails(
        mailbox=mailbox,
        criteria=criteria,
        set_flags=None,
        max=max_emails,
        low_to_high=False,
    )

    logger.info(f"Fetched {len(emails)} email(s), {len(failed)} failed to fetch.")

    extensions = tuple(ext.lower() for ext in extensions)
    if isinstance(filename_prefixes, str):
        filename_prefixes = (filename_prefixes,)
    prefixes = (
        tuple(p.lower() for p in filename_prefixes)
        if filename_prefixes
        else None
    )

    for msg in emails:
        uid: bytes = getattr(msg, "uid", None)

        for part in msg.iter_attachments():
            filename = part.get_filename() or ""
            filename_lc = filename.lower()

            if not filename_lc.endswith(extensions):
                continue

            if prefixes and not any(
                filename_lc.startswith(prefix)
                for prefix in prefixes
            ):
                continue

            content = part.get_payload(decode=True)
            if content:
                return uid, filename, content

    return None


def _coerce_to_list(value: Any) -> list[Any]:
    """
    Coerce an optional scalar/list value into a list.

    :param value: A single value, list of values, or None.
    :return: Empty list when None, original list when already a list, otherwise a single-item list.
    """
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def _normalize_iso_date(value: Any) -> str | None:
    """
    Normalize an ISO date string to YYYY-MM-DD format.

    :param value: ISO date string or None.
    :return: Normalized date string or None if input is invalid.
    """
    if not value:
        return None
    return str(value).split("T")[0]


def iter_blanket_tree_dfs(blanket_node: dict[str, Any]) -> Iterator[dict[str, Any]]:
    """
    Depth-first traversal of blanket nodes via UnderAngivelseSamling children.

    :param blanket_node: Root blanket node.
    """
    yield blanket_node
    for child in _coerce_to_list(value=blanket_node.get("UnderAngivelseSamling")):
        yield from iter_blanket_tree_dfs(child)


def _extract_felt_values_by_id(blanket: dict[str, Any]) -> dict[str, Any]:
    """
    Extract field values keyed by BlanketFeltNummerIdentifikator.

    :param blanket: Blanket node containing AngivelseFeltSamling.
    :return: Mapping {field_id -> field_value}.
    """
    values_by_id = {}
    for felt in _coerce_to_list(value=blanket.get("AngivelseFeltSamling")):
        felt_id = (
            (felt.get("BlanketFeltEnhedStruktur") or {})
            .get("BlanketFeltNummerIdentifikator")
        )
        if felt_id:
            values_by_id[str(felt_id)] = felt.get("AngivelseFeltIndholdTekst")
    return values_by_id


def _extract_rows(response: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Extract rows from SKAT Serviceplatform payload.

    The function traverses person -> income entries -> periods -> blanket tree,
    merges relevant blanket field values, and returns only rows that contain
    required income fields.

    :param response: Raw response payload from IndkomstOplysningPersonHent_O.
    :return: List of normalized report rows.
    """
    if not isinstance(response, dict):
        return []

    root = response.get("IndkomstOplysningPersonHent_O", {})
    indkomst_person_uddata = root.get("IndkomstPersonUddata") or {}
    item = (indkomst_person_uddata.get("Item") or {})

    person_strukturer = _coerce_to_list(value=item.get("IndkomstOplysningPersonStruktur"))
    rows = []

    for person in person_strukturer:
        cpr = person.get("PersonCivilRegistrationIdentifier")

        for oplysning in _coerce_to_list(value=person.get("IndkomstOplysningSamling")):
            virksomhed_se = (
                ((oplysning.get("IndberetningPligtigVirksomhedStruktur") or {})
                 .get("IndberetningPligtigVirksomhed") or {})
                .get("VirksomhedSENummerIdentifikator")
            )

            for periode in _coerce_to_list(value=oplysning.get("IndkomstLoenPeriodeSamling")):
                disposition_dato = _normalize_iso_date(value=periode.get("IndkomstPersonGruppeDispositionDato"))

                periode_items = (
                    ((periode.get("AngivelsePeriodeStruktur") or {}).get("AngivelsePeriode") or {})
                    .get("Items")
                )
                periode_items = _coerce_to_list(value=periode_items)

                angivelse_periode_fra = (
                    str(periode_items[0]).split("T")[0]
                    if len(periode_items) > 0 and periode_items[0]
                    else None
                )
                angivelse_periode_til = (
                    str(periode_items[1]).split("T")[0]
                    if len(periode_items) > 1 and periode_items[1]
                    else None
                )

                if angivelse_periode_fra and angivelse_periode_til:
                    angivelse_periode = f"{angivelse_periode_fra} - {angivelse_periode_til}"
                else:
                    angivelse_periode = angivelse_periode_fra or angivelse_periode_til

                angivelse = (periode.get("AngivelseBlanketIndholdStruktur") or {})
                for angivelse_oplysning in _coerce_to_list(value=angivelse.get("AngivelseOplysningSamling")):
                    values_11000_by_id = {}
                    values_16200_by_id = {}
                    values_16202_by_id = {}
                    values_16001_list = []

                    for blanket in iter_blanket_tree_dfs(blanket_node=angivelse_oplysning):
                        blanket_nr = str(blanket.get("BlanketNummerIdentifikator"))

                        if blanket_nr == "11000":
                            values_11000_by_id.update(_extract_felt_values_by_id(blanket=blanket))
                        elif blanket_nr == "16200":
                            values_16200_by_id.update(_extract_felt_values_by_id(blanket=blanket))
                        elif blanket_nr == "16202":
                            values_16202_by_id.update(_extract_felt_values_by_id(blanket=blanket))
                        elif blanket_nr == "16001":
                            values_16001_list.append(_extract_felt_values_by_id(blanket=blanket))

                    if values_16001_list:
                        row_sources = values_16001_list
                    elif values_16202_by_id:
                        row_sources = [{}]
                    else:
                        row_sources = []

                    for source in row_sources:
                        values_by_id = {}
                        values_by_id.update(source)
                        values_by_id.update(values_11000_by_id)
                        values_by_id.update(values_16200_by_id)
                        values_by_id.update(values_16202_by_id)

                        row = {
                            "cpr": cpr,
                            "VirksomhedSENummerIdentifikator": virksomhed_se,
                            "IndkomstPersonGruppeDispositionDato": disposition_dato,
                            "AngivelsePeriode": angivelse_periode,
                        }

                        for col_name, felt_id in get_report_field_to_blanket_field_id().items():
                            row[col_name] = values_by_id.get(felt_id)

                        row["IndkomstType"] = map_indkomsttype(value=row.get("IndkomstType"))

                        has_16001_values = any(
                            row.get(col) not in (None, "")
                            for col in _get_required_fields_from_blanket_16001()
                        )
                        has_16202_values = any(
                            row.get(col) not in (None, "")
                            for col in (
                                "Optjent BruttoFerie",
                                "BruttoFeriepenge for timelønnede",
                                "BruttoFeriepenge for fratrædende funktionærer",
                            )
                        )

                        if not (has_16001_values or has_16202_values):
                            continue

                        rows.append(row)

    return rows


def extract_rows_from_serviceplatform_response(
    payload: dict[str, Any],
) -> tuple[list[dict[str, Any]], bool]:
    """
    Extract rows from payload and indicate whether any rows were found.

    :param payload: Raw Serviceplatform payload.
    :return: Tuple of rows and a boolean indicating if any rows were found.
    """
    rows: list[dict] = []

    if not isinstance(payload, dict):
        return rows, False

    rows = _extract_rows(response=payload)
    return rows, bool(rows)


def _parse_formatted_numbers_to_numeric_series(series: pd.Series) -> pd.Series:
    """
    Parse locale-formatted number strings into numeric values.

    :param series: Input series with numeric-like values.
    :return: Numeric pandas Series with NaN for unparseable values.
    """
    text = series.fillna("").astype(str).str.strip()
    has_comma = text.str.contains(",", na=False)
    text = text.where(~has_comma, text.str.replace(".", "", regex=False))
    text = text.str.replace(",", ".", regex=False)
    return pd.to_numeric(text, errors="coerce")


def _build_month_aggregate(df: pd.DataFrame, key_cols: list[str], value_cols: list[str]) -> pd.DataFrame:
    """
    Build a monthly aggregate per key columns for configured numeric fields.

    :param df: Source dataframe with key and numeric columns.
    :param key_cols: Columns used for grouping (for example cpr).
    :param value_cols: Numeric columns to parse and aggregate.
    :return: Aggregated dataframe grouped by key columns.
    """
    out = df.copy()

    for key_col in key_cols:
        out[key_col] = out.get(key_col)
        out[key_col] = out[key_col].fillna("").astype(str).str.strip()

    for value_col in value_cols:
        out[value_col] = _parse_formatted_numbers_to_numeric_series(
            series=out.get(value_col, pd.Series([None] * len(out), index=out.index))
        )

    return (
        out.groupby(key_cols, dropna=False)[value_cols]
        .sum(min_count=1)
        .reset_index()
    )


def build_diff_table(df_prev: pd.DataFrame, df_curr: pd.DataFrame, prev_month: str, curr_month: str) -> pd.DataFrame:
    """
    Build month-to-month diff table for configured numeric fields.

    :param df_prev: Dataframe for previous month.
    :param df_curr: Dataframe for current month.
    :param prev_month: Previous month in YYYYMM.
    :param curr_month: Current month in YYYYMM.
    :return: Diff dataframe with indicators and absolute change.
    """
    def indicator(x: float) -> str:
        """
        Return an indicator symbol based on the numeric change.

        :param x: Numeric change value.
        :return: Indicator symbol ("▲" for positive, "▼" for negative, "-" for no change).
        """
        if x > 0:
            return "▲"
        if x < 0:
            return "▼"
        return "-"

    prev_agg = _build_month_aggregate(df=df_prev, key_cols=get_change_report_key_columns(), value_cols=_get_change_report_numeric_fields())
    curr_agg = _build_month_aggregate(df=df_curr, key_cols=get_change_report_key_columns(), value_cols=_get_change_report_numeric_fields())

    total_name = "Indkomstoplysning Samlet"

    prev_total = prev_agg[get_change_report_key_columns()].copy()
    prev_total[total_name] = prev_agg[_get_change_report_numeric_fields()].sum(axis=1, min_count=1)

    curr_total = curr_agg[get_change_report_key_columns()].copy()
    curr_total[total_name] = curr_agg[_get_change_report_numeric_fields()].sum(axis=1, min_count=1)

    merged = prev_total.merge(
        curr_total,
        on=get_change_report_key_columns(),
        how="outer",
        suffixes=(" (sidste)", " (denne)"),
    )

    diff = merged[get_change_report_key_columns()].copy()
    diff["FraMåned"] = prev_month
    diff["TilMåned"] = curr_month
    diff["Felt"] = total_name
    diff["Sidste måned"] = merged[f"{total_name} (sidste)"].fillna(0)
    diff["Denne måned"] = merged[f"{total_name} (denne)"].fillna(0)
    diff["Ændring"] = diff["Denne måned"] - diff["Sidste måned"]
    diff["Indikator"] = diff["Ændring"].apply(indicator)

    return (
        diff[
            [
                "FraMåned",
                "TilMåned",
                *get_change_report_key_columns(),
                "Felt",
                "Sidste måned",
                "Denne måned",
                "Indikator",
                "Ændring",
            ]
        ]
        .sort_values(["FraMåned", "TilMåned", *get_change_report_key_columns()])
        .reset_index(drop=True)
    )


def _autosize_columns(ws, df: pd.DataFrame) -> None:
    """
    Auto-size worksheet columns based on content length.

    :param ws: OpenPyXL worksheet object.
    :param df: Dataframe written to the worksheet.
    :return: None.
    """
    padding = 2
    max_width = 60

    for col_idx, col_name in enumerate(df.columns, start=1):
        series = df[col_name].fillna("").astype(str)
        max_len_raw = series.map(len).max() if not series.empty else 0
        max_len = int(max_len_raw) if pd.notna(max_len_raw) else 0
        max_len = max(max_len, len(str(col_name)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + padding, max_width)


def iter_months(start_yyyymm: str, end_yyyymm: str) -> list[str]:
    """
    Generate inclusive month range from start to end.

    :param start_yyyymm: Start month in YYYYMM.
    :param end_yyyymm: End month in YYYYMM.
    :return: List of months in YYYYMM format.
    """
    y, m = int(start_yyyymm[:4]), int(start_yyyymm[4:])
    y2, m2 = int(end_yyyymm[:4]), int(end_yyyymm[4:])

    out: list[str] = []
    while (y, m) <= (y2, m2):
        out.append(f"{y:04d}{m:02d}")
        m += 1
        if m == 13:
            y += 1
            m = 1

    return out


def write_arbejdsfortjeneste_report_excel_bytes(indkomst_df: pd.DataFrame, diff_df: pd.DataFrame) -> bytes:
    """
    Write Arbejdsfortjeneste report to Excel bytes with two sheets.

    Sheet 1: Indkomstoplysninger.
    Sheet 2: Ændring (filtered to non-zero changes when available).

    :param indkomst_df: Detailed income dataframe.
    :param diff_df: Month-to-month diff dataframe.
    :return: Excel file as bytes.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        indkomst_df.to_excel(writer, index=False, sheet_name="Indkomstoplysninger")
        ws_indkomst = writer.sheets["Indkomstoplysninger"]
        ws_indkomst.freeze_panes = "A2"
        ws_indkomst.auto_filter.ref = ws_indkomst.dimensions
        _autosize_columns(ws=ws_indkomst, df=indkomst_df)

        changes_df = diff_df.loc[diff_df["Ændring"] != 0].copy() if "Ændring" in diff_df.columns else diff_df.copy()
        changes_df.to_excel(writer, index=False, sheet_name="Ændring")
        ws_changes = writer.sheets["Ændring"]
        ws_changes.freeze_panes = "A2"
        ws_changes.auto_filter.ref = ws_changes.dimensions
        _autosize_columns(ws=ws_changes, df=changes_df)

        if len(changes_df) > 0 and "Ændring" in changes_df.columns:
            col_idx = changes_df.columns.get_loc("Ændring") + 1
            for row_idx in range(2, len(changes_df) + 2):
                cell = ws_changes.cell(row=row_idx, column=col_idx)
                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue

                if value > 0:
                    cell.font = Font(color="006100")
                elif value < 0:
                    cell.font = Font(color="9C0006")
                else:
                    cell.font = Font(color="666666")

    return output.getvalue()
