import logging
import json
import pandas as pd
import math
from io import BytesIO
from typing import Any, Sequence
from sqlalchemy import bindparam, text
from sqlalchemy.engine import Engine
from openpyxl import Workbook
from airflow.providers.http.hooks.http import HttpHook
from openpyxl.worksheet.worksheet import Worksheet
from collections import defaultdict
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from airflow.models import Variable


logger = logging.getLogger(__name__)


MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "Maj", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"]


def _get_affald_cfg() -> dict[str, Any]:
    cfg = Variable.get("affald_report_config", default_var=None, deserialize_json=True)
    if not isinstance(cfg, dict):
        raise ValueError("Airflow Variable is missing or invalid: 'affald_report_config' (expected JSON object).")
    return cfg


def _get_sheet_specs() -> list[dict[str, Any]]:
    specs = _get_affald_cfg().get("sheet_specs")
    if not isinstance(specs, list):
        raise ValueError("Missing/invalid 'sheet_specs' in Variable 'affald_report_config'.")
    return specs


def _get_default_customers() -> list[str]:
    names = _get_affald_cfg().get("genbrugspladsen_customer_names") or []
    return [str(x) for x in names]


def _normalize_name(value: Any) -> str:
    """Normalize names from SQL/config for stable matching.

    Collapses repeated whitespace and trims the ends.
    """
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def _to_float(value: Any, default: float = 1.0) -> float:
    """Parse numeric value allowing both comma and dot decimal separators."""
    if value is None:
        return default
    try:
        parsed = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return default
    return parsed if math.isfinite(parsed) else default


def sheet_specs_requires_carrier() -> bool:
    """
    Check whether any entry in ``sheet_specs`` in affald_report_config variable requires carrier (CarrierName) data.

    :return: True if carrier is needed (grouping/filtering uses carrier), otherwise False.
    """
    def _options_from_spec(spec: Any) -> dict[str, Any]:
        if isinstance(spec, dict):
            return spec
        if len(spec) >= 1 and isinstance(spec[-1], dict):
            return spec[-1]
        return {}

    for spec in _get_sheet_specs():
        options = _options_from_spec(spec=spec)

        if bool(options.get("group_by_carrier", False)):
            return True

        # sheet-level carrier filter
        if options.get("carrier_names"):
            return True

        # carrier filter inside material_groups
        mg = options.get("material_groups") or []
        for g in mg:
            if isinstance(g, dict) and g.get("carrier_names"):
                return True

    logger.debug("Carrier not needed based on sheet_specs in affald_report_config variable.")
    return False


def _default_articles_from_sheet_specs() -> list[str]:
    """
    Collect the default set of article numbers from ``sheet_specs`` in affald_report_config variable.

    :return: Sorted list of unique article numbers.
    """
    articles: set[str] = set()

    for spec in _get_sheet_specs():

        # explicit articles
        for a in spec.get("articles") or []:
            articles.add(str(a))

        # material_groups articles (for specs that omit "articles")
        for g in spec.get("material_groups") or []:
            if isinstance(g, dict):
                for a in g.get("articles") or []:
                    articles.add(str(a))

    return sorted(articles)


def fetch_affald_registration_monthly_df(
    affald_engine: Engine,
    from_date: str = "2019-01-01",
    customer_names: Sequence[str] | None = None,
    article_numbers: Sequence[str] | None = None,
    include_carrier: bool = False,
    carrier_names: Sequence[str] | None = None,
    chunksize: int = 20_000,
) -> pd.DataFrame:
    """
    Fetch monthly aggregated registration weights from ScanXNET.

    :param affald_engine: SQLAlchemy Engine for ScanXNET.
    :param from_date: Lower bound for r.DateFirst (YYYY-MM-DD).
    :param customer_names: None=default list, []=no filter, otherwise IN-filter.
    :param article_numbers: None=default list, []=no filter, otherwise IN-filter.
    :param include_carrier: If True, include CarrierName in grouping/output.
    :param carrier_names: Optional carrier IN-filter (only used when include_carrier=True).
    :param chunksize: Pandas read_sql chunksize.
    :return: DataFrame with CustomerName, ArticleNumber, year_month, weightnet_sum
             (+ CarrierName if include_carrier=True).
    """

    default_customers = _get_default_customers()
    default_articles = _default_articles_from_sheet_specs()

    if customer_names is None:
        customer_names_list: list[str] | None = list(default_customers)
    else:
        customer_names_list = list(customer_names)
        if len(customer_names_list) == 0:
            customer_names_list = None

    if article_numbers is None:
        article_numbers_list: list[str] | None = list(default_articles)
    else:
        article_numbers_list = list(article_numbers)
        if len(article_numbers_list) == 0:
            article_numbers_list = None

    carrier_names_list: list[str] | None = None
    if carrier_names is not None:
        tmp = [str(x) for x in carrier_names if str(x).strip() != ""]
        if tmp:
            carrier_names_list = tmp

    where_clauses = ["r.DateFirst >= :from_date"]
    bind_params: list[Any] = []
    params: dict[str, Any] = {"from_date": from_date}

    if customer_names_list is not None:
        where_clauses.append("r.CustomerName IN :customer_names")
        bind_params.append(bindparam("customer_names", expanding=True))
        params["customer_names"] = customer_names_list

    if article_numbers_list is not None:
        where_clauses.append("r.ArticleNumber IN :article_numbers")
        bind_params.append(bindparam("article_numbers", expanding=True))
        params["article_numbers"] = article_numbers_list

    carrier_expr = "COALESCE(NULLIF(LTRIM(RTRIM(r.CarrierName)), ''), 'Ukendt')"

    if include_carrier and carrier_names_list is not None:
        where_clauses.append(f"{carrier_expr} IN :carrier_names")
        bind_params.append(bindparam("carrier_names", expanding=True))
        params["carrier_names"] = carrier_names_list

    select_cols = ["r.CustomerName", "r.ArticleNumber"]
    group_cols = ["r.CustomerName", "r.ArticleNumber"]
    order_cols = ["r.CustomerName", "r.ArticleNumber"]

    if include_carrier:
        select_cols.append(f"{carrier_expr} AS CarrierName")
        group_cols.append(carrier_expr)
        order_cols.append("CarrierName")

    select_cols.append("CONVERT(char(7), r.DateFirst, 120) AS year_month")
    select_cols.append("SUM(CAST(r.WeightNet AS float)) AS weightnet_sum")
    group_cols.append("CONVERT(char(7), r.DateFirst, 120)")
    order_cols.append("year_month")

    stmt = text(
        f"""
        SELECT
            {", ".join(select_cols)}
        FROM ScanXNET.dbo.Registration AS r
        WHERE {" AND ".join(where_clauses)}
        GROUP BY
            {", ".join(group_cols)}
        ORDER BY
            {", ".join(order_cols)}
        """
    )

    if bind_params:
        stmt = stmt.bindparams(*bind_params)

    parts: list[pd.DataFrame] = []
    for affald_df in pd.read_sql(sql=stmt, con=affald_engine, params=params, chunksize=chunksize):
        if affald_df is not None and not affald_df.empty:
            parts.append(affald_df)

    if not parts:
        base_cols = ["CustomerName", "ArticleNumber", "year_month", "weightnet_sum"]
        if include_carrier:
            base_cols.insert(2, "CarrierName")
        return pd.DataFrame(columns=base_cols)

    logger.info(f"Fetched total {sum(len(p) for p in parts)} records from SQL for from_date={from_date}.")
    return pd.concat(parts, ignore_index=True)


def _apply_customer_and_material_view(
    df: pd.DataFrame,
    customer_names: Sequence[str] | None = None,
    customer_label: str | None = None,
    material_groups: Sequence[dict[str, Any]] | None = None,
    carrier_names: Sequence[str] | None = None,
    carrier_label: str | None = None,
    drop_unmapped_group_rows: bool = False,
) -> pd.DataFrame:
    """
    Apply customer/carrier filters and material grouping for Excel output.

    :param df: Input DataFrame.
    :param customer_names: Optional CustomerName filter (None=no filter).
    :param customer_label: If set, collapse all customers to this label (CustomerKey).
    :param material_groups: Optional mapping from ArticleNumber(s) to MaterialKey labels.
    :param carrier_names: Optional CarrierName filter at sheet-level.
    :param carrier_label: If set, collapse all carriers to this label (CarrierKey).
    :param drop_unmapped_group_rows: Drop “rest” rows for articles mentioned in material_groups.
    :return: Transformed DataFrame with CustomerKey/MaterialKey/CarrierKey columns.
    """
    if df is None or df.empty:
        out = (df.copy() if df is not None else pd.DataFrame())
        out["CustomerKey"] = pd.Series(dtype="object")
        out["MaterialKey"] = pd.Series(dtype="object")
        out["CarrierKey"] = pd.Series(dtype="object")
        return out

    out = df.copy()
    out["CustomerName"] = out["CustomerName"].map(_normalize_name)
    out["ArticleNumber"] = out["ArticleNumber"].astype(str).str.strip()
    out["weightnet_sum"] = pd.to_numeric(out["weightnet_sum"], errors="coerce").fillna(0.0)

    if "CarrierName" not in out.columns:
        out["CarrierName"] = "Ukendt"
    out["CarrierName"] = out["CarrierName"].map(_normalize_name)

    # Customer-filter pr sheet
    if customer_names is not None:
        cust = [_normalize_name(c) for c in customer_names]
        cust = [c for c in cust if c]
        if cust:
            out = out[out["CustomerName"].isin(cust)]

    # Carrier-filter pr sheet (only if the entire sheet should be limited, not just specific material groups)
    if carrier_names is not None:
        carr = [_normalize_name(c) for c in carrier_names]
        carr = [c for c in carr if c]
        if carr:
            out = out[out["CarrierName"].isin(carr)]

    # Generic customer label (summarizes all into one group)
    if customer_label is not None and str(customer_label).strip() != "":
        out["CustomerKey"] = _normalize_name(customer_label)
    else:
        out["CustomerKey"] = out["CustomerName"]

    # Generic carrier label (summarizes all into one group)
    if carrier_label is not None and str(carrier_label).strip() != "":
        out["CarrierKey"] = _normalize_name(carrier_label)
    else:
        out["CarrierKey"] = out["CarrierName"]

    # MaterialKey: default ArticleNumber, but can be grouped by material_groups config
    out["MaterialKey"] = out["ArticleNumber"]

    grouped_articles: set[str] = set()

    if material_groups:
        for g in material_groups:
            articles = [str(a).strip() for a in g.get("articles", [])]
            label = str(g.get("label", "")).strip()
            if not articles or not label:
                continue

            grouped_articles.update(articles)

            mask = out["ArticleNumber"].isin(articles)

            # customer_names filter pr material_group
            group_customers = g.get("customer_names", None)
            if group_customers is not None:
                if isinstance(group_customers, str):
                    gcust = [group_customers]
                else:
                    gcust = list(group_customers)

                gcust = [_normalize_name(c) for c in gcust]
                gcust = [c for c in gcust if c]
                if gcust:
                    mask = mask & out["CustomerName"].isin(gcust)

            # carrier_names filter pr material_group
            group_carriers = g.get("carrier_names", None)
            if group_carriers is not None:
                gc = [_normalize_name(c) for c in group_carriers]
                gc = [c for c in gc if c]
                if gc:
                    mask = mask & out["CarrierName"].isin(gc)

            # Optional per-article weighting inside this material group.
            article_weights_raw = g.get("article_weights") or {}
            if isinstance(article_weights_raw, dict) and article_weights_raw:
                article_weights: dict[str, float] = {
                    str(k).strip(): _to_float(v, default=1.0)
                    for k, v in article_weights_raw.items()
                }

                weighted_articles = set(article_weights.keys())
                weight_mask = mask & out["ArticleNumber"].isin(weighted_articles)

                if weight_mask.any():
                    factors = out.loc[weight_mask, "ArticleNumber"].map(article_weights).fillna(1.0)
                    out.loc[weight_mask, "weightnet_sum"] = (
                        out.loc[weight_mask, "weightnet_sum"].astype(float) * factors.astype(float)
                    )

            out.loc[mask, "MaterialKey"] = label

    # Remove “rest” rows that would otherwise appear as “Randers YYYY - xx.
    if drop_unmapped_group_rows and grouped_articles:
        in_grouped_article = out["ArticleNumber"].isin(sorted(grouped_articles))
        not_mapped = out["MaterialKey"].eq(out["ArticleNumber"])
        out = out[~(in_grouped_article & not_mapped)]

    logger.debug(f"After applying customer/material view: {len(out)} records (grouped articles: {sorted(grouped_articles)})")
    return out


def _pivot_material(
    df: pd.DataFrame,
    row_label_mode: str = "legacy",  # "legacy" or "generic"
    sort_mode: str = "year_then_material",  # "year_then_material" or "material_then_year"
    group_by_carrier: bool = False,
    customer_order: Sequence[str] | None = None,
    material_order: Sequence[str] | None = None,
) -> pd.DataFrame:
    """
    Pivot monthly weights into 12 month-columns + YearTotal for Excel tables.

    :param df: Input DataFrame with year_month and weightnet_sum (and *Key columns if present).
    :param row_label_mode: 'legacy' or 'generic' formatting for RowLabel.
    :param sort_mode: 'year_then_material' or 'material_then_year'.
    :param group_by_carrier: If True, include CarrierKey in pivot index/output.
    :param customer_order: Optional explicit order for CustomerKey (list of customer names).
                           Customers not listed will be appended after the listed ones.
    :return: Pivoted DataFrame with columns for months 1..12 and YearTotal.
    """
    base_cols = ["CustomerKey", "year", "MaterialKey"]
    if group_by_carrier:
        base_cols.append("CarrierKey")
    base_cols.append("RowLabel")

    if df is None or df.empty:
        return pd.DataFrame(columns=base_cols + list(range(1, 13)) + ["YearTotal"])

    tmp = df.copy()
    tmp["year_month"] = tmp["year_month"].astype(str)
    tmp["year"] = tmp["year_month"].str.slice(0, 4).astype(int)
    tmp["month"] = tmp["year_month"].str.slice(5, 7).astype(int)

    if "CustomerKey" not in tmp.columns:
        tmp["CustomerKey"] = tmp["CustomerName"].astype(str)
    if "MaterialKey" not in tmp.columns:
        tmp["MaterialKey"] = tmp["ArticleNumber"].astype(str)
    if group_by_carrier and "CarrierKey" not in tmp.columns:
        if "CarrierName" in tmp.columns:
            tmp["CarrierKey"] = tmp["CarrierName"].astype(str)
        else:
            tmp["CarrierKey"] = "Ukendt"

    idx = ["CustomerKey", "year", "MaterialKey"] + (["CarrierKey"] if group_by_carrier else [])

    pivot = tmp.pivot_table(
        index=idx,
        columns="month",
        values="weightnet_sum",
        aggfunc="sum",
    )

    for m in range(1, 13):
        if m not in pivot.columns:
            pivot[m] = 0.0

    pivot = pivot[[m for m in range(1, 13)]]
    pivot["YearTotal"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    # Apply customer ordering
    if customer_order:
        order = [_normalize_name(x) for x in customer_order]
        order = [x for x in order if x]

        present = pivot["CustomerKey"].astype(str).tolist()
        present_unique = list(dict.fromkeys(present))  # keep first-seen order
        present_set = set(present_unique)

        known = [x for x in order if x in present_set]
        known_set = set(known)
        unknown = [x for x in present_unique if x not in known_set]
        cats = known + unknown

        pivot["CustomerKey"] = pd.Categorical(pivot["CustomerKey"], categories=cats, ordered=True)

    # Apply material ordering
    if material_order:
        order = [_normalize_name(x) for x in material_order]
        order = [x for x in order if x]

        present = pivot["MaterialKey"].astype(str).tolist()
        present_unique = list(dict.fromkeys(present))
        present_set = set(present_unique)

        known = [x for x in order if x in present_set]
        known_set = set(known)
        unknown = [x for x in present_unique if x not in known_set]
        cats = known + unknown

        pivot["MaterialKey"] = pd.Categorical(
            pivot["MaterialKey"],
            categories=cats,
            ordered=True
        )

    if group_by_carrier:
        if sort_mode == "material_then_year":
            pivot = pivot.sort_values(["CustomerKey", "MaterialKey", "CarrierKey", "year"], kind="stable")
        else:
            pivot = pivot.sort_values(["CustomerKey", "year", "MaterialKey", "CarrierKey"], kind="stable")
    else:
        if sort_mode == "material_then_year":
            pivot = pivot.sort_values(["CustomerKey", "MaterialKey", "year"], kind="stable")
        else:
            pivot = pivot.sort_values(["CustomerKey", "year", "MaterialKey"], kind="stable")

    def _mk_rowlabel(r: pd.Series) -> str:
        cust = str(r["CustomerKey"])
        year = str(int(r["year"]))
        mat = str(r["MaterialKey"])
        carr = str(r["CarrierKey"]) if group_by_carrier else ""

        if row_label_mode == "legacy":
            base = f"{cust} {year} - (vare nr. {mat})"
        else:
            base = f"{cust} {year} - {mat}"

        if group_by_carrier:
            return f"{base} - {carr}"
        return base

    pivot["RowLabel"] = pivot.apply(_mk_rowlabel, axis=1)

    out_cols = ["CustomerKey", "year", "MaterialKey"] + (["CarrierKey"] if group_by_carrier else []) + ["RowLabel"]
    return pivot[out_cols + [m for m in range(1, 13)] + ["YearTotal"]]


def _write_sheet(
    ws: Worksheet,
    table: pd.DataFrame,
    title: str,
    ton_label_mode: str = "legacy",
    blank_between: str = "customer",
    group_by_carrier: bool = False,
    material_order: Sequence[str] | None = None,
    ytd_year: int | None = None,
    ytd_end_month: int | None = None,
) -> None:
    """
    Write a pivoted table (from ``_pivot_material``) into an Excel worksheet.

    Adds an extra YTD % column for the newest year (e.g. 2026), comparing Jan..M vs same Jan..M last year.
    Keeps the existing annual % column unchanged.
    """
    end_m: int | None = None
    if ytd_end_month is not None:
        try:
            end_m = int(ytd_end_month)
        except Exception:
            end_m = None
    if end_m is not None:
        end_m = max(1, min(12, end_m))

    ytd_caption = "YTD"
    if end_m is not None:
        ytd_caption = f"YTD (Jan-{MONTH_NAMES[end_m - 1]})"

    # A..P (1 label + 12 months + 1 year total + 1 pct + 1 ytd pct)
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("N2:P2")
    ws["N2"].value = "Årssummer"
    ws["N2"].font = Font(bold=True)

    ws["A3"].value = "i kg."
    ws["A3"].font = Font(bold=True)

    ws["N3"].value = "i kg"
    ws["N3"].font = Font(bold=True)
    ws["O3"].value = "i %"
    ws["O3"].font = Font(bold=True)
    ws["P3"].value = "i %"
    ws["P3"].font = Font(bold=True)

    # Header (row 4)
    for i, name in enumerate(MONTH_NAMES, start=2):  # B..M
        cell = ws.cell(row=4, column=i, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws["N4"].value = "Årssum"
    ws["N4"].font = Font(bold=True)
    ws["N4"].alignment = Alignment(horizontal="center")

    ws["O4"].value = "Stigning = +/Fald = -"
    ws["O4"].font = Font(bold=True)
    ws["O4"].alignment = Alignment(horizontal="center")

    ws["P4"].value = ytd_caption
    ws["P4"].font = Font(bold=True)
    ws["P4"].alignment = Alignment(horizontal="center")

    num_fmt_kg = "#,##0"
    num_fmt_ton = "#,##0.00"
    num_fmt_pct = "+0.0%;-0.0%;0.0%"
    current_year_fill = PatternFill(fill_type="solid", start_color="FFD9E1F2", end_color="FFD9E1F2")

    if table is None or table.empty:
        ws.column_dimensions["A"].width = 67
        for col in range(2, 17):  # B..P
            ws.column_dimensions[get_column_letter(col)].width = 20
        return

    # KG-section
    start_row = 5
    excel_row = start_row
    prev_customer: str | None = None
    prev_material: str | None = None

    # pct change pr. key (annual)
    prev_year_total_by_key: dict[tuple[str, str, str] | tuple[str, str], float] = {}

    records = table.to_dict(orient="records")

    # Precompute YTD totals per (key, year) for KG section
    ytd_total_by_key_year: dict[tuple[Any, int], float] = {}
    if ytd_year is not None and end_m is not None:
        for rec in records:
            year_val = rec.get("year", None)
            if year_val is None:
                continue
            year_int = int(year_val)

            customer = str(rec.get("CustomerKey", ""))
            material = str(rec.get("MaterialKey", ""))
            carrier = str(rec.get("CarrierKey", "")) if group_by_carrier else ""
            key = (customer, material, carrier) if group_by_carrier else (customer, material)

            ytd_sum = 0.0
            for m in range(1, end_m + 1):
                v = rec.get(m, 0.0)
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    v = 0.0
                ytd_sum += float(v)
            ytd_total_by_key_year[(key, year_int)] = float(ytd_sum)

    for rec in records:
        year_val = rec.get("year", None)
        year_int = int(year_val) if year_val is not None else None
        customer = str(rec.get("CustomerKey", ""))
        material = str(rec.get("MaterialKey", ""))
        carrier = str(rec.get("CarrierKey", "")) if group_by_carrier else ""
        row_label = rec.get("RowLabel", "")

        insert_blank = False
        if blank_between in ("customer", "customer_and_material"):
            if prev_customer is not None and customer != prev_customer:
                insert_blank = True
        if blank_between in ("material", "customer_and_material"):
            if prev_material is not None and material != prev_material:
                insert_blank = True
        if group_by_carrier and blank_between in ("material", "customer_and_material", "none"):
            pass

        if insert_blank:
            excel_row += 1

        ws.cell(row=excel_row, column=1, value=row_label)

        for m in range(1, 13):
            val = rec.get(m, 0.0)
            if pd.isna(val):
                val = 0.0
            cell = ws.cell(row=excel_row, column=1 + m, value=float(val))
            cell.number_format = num_fmt_kg

            if ytd_year is not None and year_int == int(ytd_year):
                cell.fill = current_year_fill

        y_val = rec.get("YearTotal", 0.0)
        if pd.isna(y_val):
            y_val = 0.0
        y_total = float(y_val)

        y = ws.cell(row=excel_row, column=14, value=y_total)  # N
        y.number_format = num_fmt_kg
        y.font = Font(bold=True)

        key = (customer, material, carrier) if group_by_carrier else (customer, material)

        # Annual % (existing behavior) -> O
        prev_total = prev_year_total_by_key.get(key)
        if prev_total is None or prev_total == 0:
            ws.cell(row=excel_row, column=15, value=None)  # O
        else:
            pct = (y_total - prev_total) / prev_total
            cell_pct = ws.cell(row=excel_row, column=15, value=float(pct))
            cell_pct.number_format = num_fmt_pct

        prev_year_total_by_key[key] = y_total

        # YTD % (only for newest year) -> P
        ytd_cell = ws.cell(row=excel_row, column=16, value=None)  # P
        if ytd_year is not None and end_m is not None:
            year_val = rec.get("year", None)
            if year_val is not None and int(year_val) == int(ytd_year):
                curr_ytd = ytd_total_by_key_year.get((key, int(ytd_year)))
                prev_ytd = ytd_total_by_key_year.get((key, int(ytd_year) - 1))
                if prev_ytd is None or prev_ytd == 0 or curr_ytd is None:
                    ytd_cell.value = None
                else:
                    ytd_pct = (float(curr_ytd) - float(prev_ytd)) / float(prev_ytd)
                    ytd_cell.value = float(ytd_pct)
                    ytd_cell.number_format = num_fmt_pct

        prev_customer = customer
        prev_material = material
        excel_row += 1

    last_kg_row = excel_row - 1

    # TON-section
    raw_records: list[dict[str | int, Any]] = []
    for rec in records:
        year = rec.get("year", None)
        material = rec.get("MaterialKey", None)
        if year is None or material is None:
            continue

        rr: dict[str | int, Any] = {"year": int(year), "MaterialKey": str(material)}
        if group_by_carrier:
            rr["CarrierKey"] = str(rec.get("CarrierKey", ""))

        for m in range(1, 13):
            val = rec.get(m, 0.0)
            rr[m] = float(0.0 if pd.isna(val) else val)

        y_val = rec.get("YearTotal", 0.0)
        rr["YearTotal"] = float(0.0 if pd.isna(y_val) else y_val)
        raw_records.append(rr)

    if raw_records:
        months = list(range(1, 13))
        ton_df = pd.DataFrame.from_records(raw_records)

        grp_cols = ["year", "MaterialKey"] + (["CarrierKey"] if group_by_carrier else [])
        ton_df = (
            ton_df.groupby(grp_cols, as_index=False)[months + ["YearTotal"]]
            .sum()
            .sort_values(grp_cols, kind="stable")
        )

        ton_df[months + ["YearTotal"]] = ton_df[months + ["YearTotal"]] / 1000.0

        # Apply material ordering for TON-section
        if material_order:
            order = [_normalize_name(x) for x in material_order]
            order = [x for x in order if x]

            present = ton_df["MaterialKey"].astype(str).tolist()
            present_unique = list(dict.fromkeys(present))
            present_set = set(present_unique)

            known = [x for x in order if x in present_set]
            known_set = set(known)
            unknown = [x for x in present_unique if x not in known_set]
            cats = known + unknown

            ton_df["MaterialKey"] = pd.Categorical(
                ton_df["MaterialKey"],
                categories=cats,
                ordered=True,
            )

        # Compute YTD totals (tons)
        if end_m is not None:
            ytd_month_cols = [m for m in range(1, end_m + 1)]
            ton_df["YTDTotal"] = ton_df[ytd_month_cols].sum(axis=1)
        else:
            ton_df["YTDTotal"] = 0.0

        # pct change year-over-year (annual + ytd)
        if group_by_carrier:
            ton_df = ton_df.sort_values(["MaterialKey", "CarrierKey", "year"], kind="stable")
            ton_df["PctChange"] = ton_df.groupby(["MaterialKey", "CarrierKey"])["YearTotal"].pct_change()
            ton_df["YTDChange"] = ton_df.groupby(["MaterialKey", "CarrierKey"])["YTDTotal"].pct_change()
            ton_df = ton_df.sort_values(["year", "MaterialKey", "CarrierKey"], kind="stable")
        else:
            ton_df = ton_df.sort_values(["MaterialKey", "year"], kind="stable")
            ton_df["PctChange"] = ton_df.groupby("MaterialKey")["YearTotal"].pct_change()
            ton_df["YTDChange"] = ton_df.groupby("MaterialKey")["YTDTotal"].pct_change()
            ton_df = ton_df.sort_values(["year", "MaterialKey"], kind="stable")

        def _ton_label(r: pd.Series) -> str:
            y = int(r["year"])
            mkey = str(r["MaterialKey"])
            if ton_label_mode == "legacy":
                base = f"{y} - varenummer {mkey} (total i tons)"
            else:
                base = f"{y} {mkey}"
            if group_by_carrier:
                return f"{base} - {str(r.get('CarrierKey', ''))}"
            return base

        ton_df["RowLabel"] = ton_df.apply(_ton_label, axis=1)

        ton_unit_row = last_kg_row + 2
        ton_header_row = ton_unit_row + 1
        ton_data_row = ton_header_row + 1

        ws.cell(row=ton_unit_row, column=1, value="i tons.")
        ws.cell(row=ton_unit_row, column=1).font = Font(bold=True)

        ws.cell(row=ton_unit_row, column=14, value="i tons")
        ws.cell(row=ton_unit_row, column=14).font = Font(bold=True)

        ws.cell(row=ton_unit_row, column=15, value="i %")
        ws.cell(row=ton_unit_row, column=15).font = Font(bold=True)

        ws.cell(row=ton_unit_row, column=16, value="i %")
        ws.cell(row=ton_unit_row, column=16).font = Font(bold=True)

        for i, name in enumerate(MONTH_NAMES, start=2):  # B..M
            cell = ws.cell(row=ton_header_row, column=i, value=name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        n_cell = ws.cell(row=ton_header_row, column=14, value="Årssum")
        n_cell.font = Font(bold=True)
        n_cell.alignment = Alignment(horizontal="center")

        o_cell = ws.cell(row=ton_header_row, column=15, value="Stigning = +/Fald = -")
        o_cell.font = Font(bold=True)
        o_cell.alignment = Alignment(horizontal="center")

        p_cell = ws.cell(row=ton_header_row, column=16, value=ytd_caption)
        p_cell.font = Font(bold=True)
        p_cell.alignment = Alignment(horizontal="center")

        r = ton_data_row
        prev_year: int | None = None
        for rec in ton_df.to_dict(orient="records"):
            year_int = int(rec.get("year"))
            if prev_year is not None and year_int != prev_year:
                r += 1

            ws.cell(row=r, column=1, value=rec.get("RowLabel", ""))

            for m in range(1, 13):
                val = rec.get(m, 0.0)
                if pd.isna(val):
                    val = 0.0
                cell = ws.cell(row=r, column=1 + m, value=float(val))
                cell.number_format = num_fmt_ton

                if ytd_year is not None and year_int == int(ytd_year):
                    cell.fill = current_year_fill

            y_val = rec.get("YearTotal", 0.0)
            if pd.isna(y_val):
                y_val = 0.0
            y = ws.cell(row=r, column=14, value=float(y_val))
            y.number_format = num_fmt_ton
            y.font = Font(bold=True)

            pct_val = rec.get("PctChange", None)
            if pct_val is None or (isinstance(pct_val, float) and pd.isna(pct_val)):
                ws.cell(row=r, column=15, value=None)
            else:
                cell_pct = ws.cell(row=r, column=15, value=float(pct_val))
                cell_pct.number_format = num_fmt_pct

            # YTD % only for newest year
            ytd_val = rec.get("YTDChange", None)
            if ytd_year is None or year_int != int(ytd_year):
                ws.cell(row=r, column=16, value=None)
            else:
                if ytd_val is None or (isinstance(ytd_val, float) and pd.isna(ytd_val)):
                    ws.cell(row=r, column=16, value=None)
                else:
                    cell_ytd = ws.cell(row=r, column=16, value=float(ytd_val))
                    cell_ytd.number_format = num_fmt_pct

            prev_year = year_int
            r += 1

    ws.column_dimensions["A"].width = 67
    for col in range(2, 17):  # B..P
        ws.column_dimensions[get_column_letter(col)].width = 20


def _append_vare_nr_if_missing(label: str, articles: Sequence[str]) -> str:
    """
    Append '(vare nr. ...)' to a label if it is not already present.

    :param label: Base label.
    :param articles: Article numbers to include (duplicates/blank values are ignored).
    :return: Label with appended article numbers (or unchanged label).
    """
    lbl = (label or "").strip()
    if not lbl:
        return lbl

    if "vare nr" in lbl.lower():
        return lbl

    # Remaining order, remove duplicates/blank values
    seen: set[str] = set()
    cleaned: list[str] = []
    for a in articles or []:
        s = str(a).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        cleaned.append(s)

    if not cleaned:
        return lbl

    return f"{lbl} (vare nr. {', '.join(cleaned)})"


def _normalize_material_groups_for_sheet(
    material_groups: Sequence[dict[str, Any]] | None,
    auto_append_vare_nr: bool,
) -> list[dict[str, Any]] | None:
    """
    Normalize material_groups for a sheet (optionally auto-append vare nr. to labels).

    :param material_groups: Material group config from ``sheet_specs`` in affald_report_config variable.
    :param auto_append_vare_nr: Default behaviour for auto-appending vare nr. (can be overridden per group).
    :return: Normalized copy of material_groups (or None).
    """
    if not material_groups:
        return None

    out: list[dict[str, Any]] = []
    for i, g in enumerate(material_groups):
        if not isinstance(g, dict):
            raise TypeError(
                f"material_groups[{i}] must be a dict, got {type(g).__name__}: {g!r}"
            )

        gg = dict(g)
        group_auto = gg.get("auto_append_vare_nr", auto_append_vare_nr)
        if group_auto:
            lbl = str(gg.get("label", "")).strip()
            arts = [str(a) for a in (gg.get("articles") or [])]
            gg["label"] = _append_vare_nr_if_missing(label=lbl, articles=arts)

        out.append(gg)

    return out


def _build_affald_excel_by_article_workbook(df: pd.DataFrame) -> Workbook:
    """
    Build the affald Excel workbook from a DataFrame and ``sheet_specs`` in affald_report_config variable.

    :param df: Input DataFrame (typically output from fetch_affald_registration_monthly_df).
    :return: openpyxl Workbook with one sheet per entry in ``sheet_specs`` in affald_report_config variable.
    """
    df2 = df.copy()
    df2["ArticleNumber"] = df2["ArticleNumber"].astype(str)
    df2["CustomerName"] = df2["CustomerName"].astype(str)
    if "CarrierName" in df2.columns:
        df2["CarrierName"] = df2["CarrierName"].astype(str)

    sheet_specs: list[Any] = _get_sheet_specs()

    wb = Workbook()
    default_ws = wb.active
    wb.remove(worksheet=default_ws)

    for spec in sheet_specs:
        sheet_name = str(spec["sheet_name"])
        title = str(spec["title"])
        row_label_mode = str(spec.get("row_label_mode", "legacy"))
        ton_label_mode = str(spec.get("ton_label_mode", "legacy"))
        sort_mode = str(spec.get("sort_mode", "year_then_material"))
        blank_between = str(spec.get("blank_between", "customer"))

        group_by_carrier = bool(spec.get("group_by_carrier", False))
        carrier_names = spec.get("carrier_names", None)
        carrier_label = spec.get("carrier_label", None)

        customer_names = spec.get("customer_names", None)
        customer_label = spec.get("customer_label", None)

        auto_append_vare_nr = bool(spec.get("auto_append_vare_nr", True))

        material_groups = spec.get("material_groups", None)
        material_groups = _normalize_material_groups_for_sheet(
            material_groups,
            auto_append_vare_nr=auto_append_vare_nr,
        )

        drop_unmapped_group_rows = bool(spec.get("drop_unmapped_group_rows", False))

        articles = spec.get("articles", None)
        if articles is None:
            articles_list: list[str] = []
            if material_groups:
                for g in material_groups:
                    articles_list.extend([str(a) for a in g.get("articles", [])])
            articles = sorted(set(articles_list))
        else:
            articles = [str(a) for a in articles]

        ws = wb.create_sheet(title=sheet_name)

        subset = df2[df2["ArticleNumber"].isin([str(a) for a in articles])]
        subset = _apply_customer_and_material_view(
            df=subset,
            customer_names=customer_names,
            customer_label=customer_label,
            material_groups=material_groups,
            carrier_names=carrier_names,
            carrier_label=carrier_label,
            drop_unmapped_group_rows=drop_unmapped_group_rows,
        )

        ytd_year: int | None = None
        ytd_end_month: int | None = None
        if subset is not None and (not subset.empty) and ("year_month" in subset.columns):
            ym = subset["year_month"].astype(str)
            years = ym.str.slice(0, 4).astype(int)
            months = ym.str.slice(5, 7).astype(int)

            if not years.empty:
                ytd_year = int(years.max())
                months_in_ytd_year = months[years == ytd_year]
                if not months_in_ytd_year.empty:
                    ytd_end_month = int(months_in_ytd_year.max())

        customer_order = spec.get("customer_order")
        if customer_order is None:
            customer_order = spec.get("customer_names")

        material_order = spec.get("material_order")

        table = _pivot_material(
            df=subset,
            row_label_mode=row_label_mode,
            sort_mode=sort_mode,
            group_by_carrier=group_by_carrier,
            customer_order=customer_order,
            material_order=material_order,
        )
        _write_sheet(
            ws=ws,
            table=table,
            title=title,
            ton_label_mode=ton_label_mode,
            blank_between=blank_between,
            group_by_carrier=group_by_carrier,
            material_order=material_order,
            ytd_year=ytd_year,
            ytd_end_month=ytd_end_month,
        )

    return wb


def build_affald_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Build the affald Excel report in memory and return it as bytes.

    :param df: Input DataFrame used to build the workbook.
    :return: Excel file content as bytes
    """
    wb = _build_affald_excel_by_article_workbook(df=df)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def mp_waste_amount_data(
    http_hook: HttpHook,
    customer_numbers: list[int],
    from_date: str,
    to_date: str,
    installation_address_id: list[int],
) -> list[dict[str, Any]]:
    """
    Fetch waste amount statistics from the MP API

    :param http_hook: Airflow HttpHook configured for the MP API base URL/connection.
    :param customer_numbers: Customer number filter sent to the API payload.
    :param from_date: Start date (string) sent to the API payload (expected format per API, e.g. YYYY-MM-DD).
    :param to_date: End date (string) sent to the API payload (expected format per API, e.g. YYYY-MM-DD).
    :param installation_address_id: Installation address id filter sent to the API payload.
    :return: List of raw row dicts returned from the API (concatenated across pages).
    """
    logger.info("Fetching waste amount data from MP API...")

    conn = http_hook.get_connection(http_hook.http_conn_id)
    if not conn.password:
        raise ValueError("Missing MP_ApiKey (connection password) for marius_pedersen_api connection.")

    headers = {
        "MP_ApiKey": conn.password,
        "Content-Type": "application/json",
    }

    http_hook.method = "POST"

    page_size = 200
    page = 1

    all_rows: list[dict[str, Any]] = []

    total_count: int | None = None
    max_pages: int | None = None

    while True:
        logger.info(f"Fetching MP data page={page} page_size={page_size} ...")

        payload = {
            "PageSize": page_size,
            "Page": page,
            "CustomerNumbers": customer_numbers,
            "FromDate": from_date,
            "ToDate": to_date,
            "InstallationAddressId": installation_address_id,
        }

        res = http_hook.run(
            endpoint="/umbraco/api/wastestatistic/GetWasteamountStatistic",
            data=json.dumps(payload),
            headers=headers,
        )

        body = res.json()
        data = body.get("data", [])

        # find totalcount on first page and compute expected max pages
        if total_count is None:
            total_count = int(body.get("totalCount", 0))
            max_pages = math.ceil(total_count / page_size) if total_count else 0

            logger.info(f"MP totalCount={total_count}, expected_pages={max_pages}")

        years = [
            row.get("activityYear")
            for row in data
            if row.get("activityYear") is not None
        ]

        months = [
            int(row.get("activityMonth"))
            for row in data
            if row.get("activityMonth") is not None
        ]

        total_so_far_after = len(all_rows) + len(data)
        logger.info(
            f"MP page={page} status={res.status_code} rows={len(data)} years={sorted(set(years)) if years else []} "
            f"months={sorted(set(months)) if months else []} total_so_far={total_so_far_after}"
        )

        # save data if the page is not empty
        if data:
            all_rows.extend(data)
        elif (total_count or 0) > 0:
            logger.warning(f"MP page={page} returned 0 rows although totalCount={total_count}")

        # stop when reaching the last expected page
        if max_pages is not None:
            if max_pages == 0:
                logger.info("MP totalCount=0; no pages to fetch.")
                break
            if page >= max_pages:
                logger.info(f"Reached last expected page {page}/{max_pages}")
                break

        page += 1

    logger.info(f"Successfully retrieved data from MP API. Total records: {len(all_rows)} (pages fetched: {page})")

    if not all_rows:
        raise ValueError(f"No data returned from MP API for customers {customer_numbers} and date range {from_date} to {to_date}.")

    return all_rows


def aggregate_taxes_quantity_by_month(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Aggregate MP ``taxesQuantity(Mængde kg)`` per (customerNumber, activityYear, activityMonth).

    :param rows: List of API rows (dict-like) containing customerNumber/activityYear/activityMonth/taxesQuantity.
    :return: List of dicts with keys: customerNumber, activityYear, activityMonth, taxesQuantity.
    """
    monthly_totals = defaultdict(float)

    for row in rows:
        if not isinstance(row, dict):
            continue

        cust = row.get("customerNumber")
        year = row.get("activityYear")
        month = row.get("activityMonth")
        quantity = row.get("taxesQuantity") or 0

        if cust is None or year is None or month is None:
            continue

        monthly_totals[(int(cust), int(year), int(month))] += float(quantity)

    return [
        {
            "customerNumber": cust,
            "activityYear": year,
            "activityMonth": month,
            "taxesQuantity": total,
        }
        for (cust, year, month), total in sorted(monthly_totals.items())
    ]


def build_mp_monthly_excel_bytes(
    monthly_data: list[dict[str, Any]],
) -> bytes:
    """
    Build the MP Excel report file for MP monthly quantities and return it as bytes.

    :param monthly_data: Monthly records containing customerNumber/activityYear/activityMonth/taxesQuantity.
                         Values are summed per (customer, year, month) before writing.
    :return: Excel file content as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Marius Pedersen udtræk"

    customer_number_name_map: dict[int, str] = {
        80067523: "MP A/S Undergrund Århus VIP",
        80070490: "Brændbart, direkte Aarhus MP (mini aff.)",
        80070170: "Pap - indsamlet i containere MP (pap)",
    }

    def _row_label(cust_i: int, year_i: int) -> str:
        name = customer_number_name_map.get(cust_i)
        if name:
            return f"{name} {year_i}"
        return f"{cust_i} {year_i}"

    # Find max-år til highlight (som i _write_sheet)
    years_present: list[int] = []
    for rec in (monthly_data or []):
        y = rec.get("activityYear")
        if y is not None:
            try:
                years_present.append(int(y))
            except Exception:
                pass
    newest_year = max(years_present) if years_present else None

    # Header/top som _write_sheet (A..P)
    ws.merge_cells("A1:P1")
    ws["A1"].value = "Marius Pedersen API - Mængder"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("N2:P2")
    ws["N2"].value = "Årssummer"
    ws["N2"].font = Font(bold=True)

    ws["A3"].value = "i kg."
    ws["A3"].font = Font(bold=True)

    ws["N3"].value = "i kg"
    ws["N3"].font = Font(bold=True)
    ws["O3"].value = "i %"
    ws["O3"].font = Font(bold=True)

    # Header-row 4: months + year total + pct change
    for i, name in enumerate(MONTH_NAMES, start=2):  # B..M
        cell = ws.cell(row=4, column=i, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws["N4"].value = "Årssum"
    ws["N4"].font = Font(bold=True)
    ws["N4"].alignment = Alignment(horizontal="center")

    ws["O4"].value = "Stigning = +/Fald = -"
    ws["O4"].font = Font(bold=True)
    ws["O4"].alignment = Alignment(horizontal="center")

    num_fmt_kg = "#,##0"
    num_fmt_ton = "#,##0.00"
    num_fmt_pct = "+0.0%;-0.0%;0.0%"
    current_year_fill = PatternFill(fill_type="solid", start_color="FFD9E1F2", end_color="FFD9E1F2")

    # Build pivot: (customer, year) -> month -> qty
    month_qty: dict[tuple[int, int], dict[int, float]] = defaultdict(lambda: defaultdict(float))

    for rec in (monthly_data or []):
        cust = rec.get("customerNumber")
        year = rec.get("activityYear")
        month = rec.get("activityMonth")
        qty = rec.get("taxesQuantity") or 0

        if cust is None or year is None or month is None:
            continue

        try:
            cust_i = int(cust)
            year_i = int(year)
            month_i = int(month)
        except Exception:
            continue

        if not (1 <= month_i <= 12):
            continue

        month_qty[(cust_i, year_i)][month_i] += float(qty)

    keys_sorted = sorted(month_qty.keys(), key=lambda t: (t[0], t[1]))

    ws.column_dimensions["A"].width = 67
    for col in range(2, 17):  # B..P
        ws.column_dimensions[get_column_letter(col)].width = 20

    def _write_customer_year_section(
        start_row: int,
        unit_divisor: float,
        number_format: str,
        prev_year_total_by_customer: dict[int, float],
    ) -> int:
        if float(unit_divisor) == 0.0:
            raise ValueError("unit_divisor must be non-zero")

        r = start_row
        prev_customer: int | None = None

        for (cust_i, year_i) in keys_sorted:
            if prev_customer is not None and cust_i != prev_customer:
                r += 1  # blank row between customers

            ws.cell(row=r, column=1, value=_row_label(cust_i, year_i))

            year_total = 0.0
            for m in range(1, 13):
                v = float(month_qty[(cust_i, year_i)].get(m, 0.0)) / float(unit_divisor)
                year_total += v

                cell = ws.cell(row=r, column=1 + m, value=float(v))  # B..M
                cell.number_format = number_format
                if newest_year is not None and year_i == newest_year:
                    cell.fill = current_year_fill

            y_cell = ws.cell(row=r, column=14, value=float(year_total))  # N
            y_cell.number_format = number_format
            y_cell.font = Font(bold=True)

            prev_total = prev_year_total_by_customer.get(cust_i)
            if prev_total is None or prev_total == 0:
                ws.cell(row=r, column=15, value=None)  # O
            else:
                pct = (year_total - prev_total) / prev_total
                p_cell = ws.cell(row=r, column=15, value=float(pct))
                p_cell.number_format = num_fmt_pct

            prev_year_total_by_customer[cust_i] = float(year_total)
            prev_customer = cust_i
            r += 1

        return r - 1

    # KG-section
    start_row = 5
    prev_year_total_by_customer_kg: dict[int, float] = {}
    last_kg_row = _write_customer_year_section(
        start_row=start_row,
        unit_divisor=1.0,
        number_format=num_fmt_kg,
        prev_year_total_by_customer=prev_year_total_by_customer_kg,
    )

    # TON-section
    ton_unit_row = last_kg_row + 2  # blank row + unit row
    ton_header_row = ton_unit_row + 1  # header row with month names, year total, pct change
    ton_data_row = ton_header_row + 1  # first data row for tons

    ws.cell(row=ton_unit_row, column=1, value="i tons.")
    ws.cell(row=ton_unit_row, column=1).font = Font(bold=True)

    ws.cell(row=ton_unit_row, column=14, value="i tons")
    ws.cell(row=ton_unit_row, column=14).font = Font(bold=True)

    ws.cell(row=ton_unit_row, column=15, value="i %")
    ws.cell(row=ton_unit_row, column=15).font = Font(bold=True)

    for i, name in enumerate(MONTH_NAMES, start=2):  # B..M
        cell = ws.cell(row=ton_header_row, column=i, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.cell(row=ton_header_row, column=14, value="Årssum").font = Font(bold=True)
    ws.cell(row=ton_header_row, column=15, value="Stigning = +/Fald = -").font = Font(bold=True)

    prev_year_total_by_customer_ton: dict[int, float] = {}
    _write_customer_year_section(
        start_row=ton_data_row,
        unit_divisor=1000.0,
        number_format=num_fmt_ton,
        prev_year_total_by_customer=prev_year_total_by_customer_ton,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
